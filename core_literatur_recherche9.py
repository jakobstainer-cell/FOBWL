
# core_literatur_recherche9.py
# Basierend auf literatur_recherche9.py – refactored für Streamlit
# Erweiterung: generischer Modus + query-basiertes Ranking + Must/Should/Exclude + Facettenfilter

import os
import re
import time
import json
import html
from io import BytesIO
from typing import List, Dict, Optional, Any, Tuple, Union
import xml.etree.ElementTree as ET

import requests
import pandas as pd
from openpyxl import load_workbook


# ============================================================
# 1) Defaults / Konfiguration (werden über config überschrieben)
# ============================================================

HTTP_TIMEOUT = 25

MAILTO = os.getenv("MAILTO", "fhb231351@fh-vie.ac.at")
USER_AGENT = f"LiteraturRechercheScript/9.1 (mailto:{MAILTO})"

OPENALEX_API_KEY = os.getenv("OPENALEX_API_KEY", "").strip()
UNPAYWALL_EMAIL = os.getenv("UNPAYWALL_EMAIL", MAILTO).strip()
SEMANTIC_SCHOLAR_API_KEY = os.getenv("SEMANTIC_SCHOLAR_API_KEY", "").strip()

# Publisher hints (Crossref)
PUBLISHER_NAME_HINTS = ["Springer", "Gabler", "Vahlen", "Hanser", "Schäffer-Poeschel"]
PUBLISHER_LOCATION_HINTS = ["Germany", "Austria"]


# ============================================================
# 2) Rate limiting & Backoff
# ============================================================

API_MIN_INTERVAL = {
    "CrossRef": 0.25,
    "OpenLibrary": 0.25,
    "OpenAlex": 0.35,
    "SemanticScholar": 3.6 if not SEMANTIC_SCHOLAR_API_KEY else 0.35,
    "DNB_SRU": 0.25,
    "Unpaywall": 0.35,
}

_last_call_ts: Dict[str, float] = {}


def throttle(api_name: str):
    min_interval = API_MIN_INTERVAL.get(api_name, 0.25)
    now = time.time()
    last = _last_call_ts.get(api_name, 0.0)
    wait = min_interval - (now - last)
    if wait > 0:
        time.sleep(wait)
    _last_call_ts[api_name] = time.time()


def request_with_backoff(
    session: requests.Session,
    method: str,
    url: str,
    *,
    api_name: str,
    params: Optional[dict] = None,
    headers: Optional[dict] = None,
    timeout: int = 25,
    max_retries: int = 6,
) -> requests.Response:
    delay = 2.0
    for _ in range(max_retries):
        throttle(api_name)
        r = session.request(method, url, params=params, headers=headers, timeout=timeout)

        if r.status_code < 400:
            return r

        if r.status_code == 429:
            retry_after = r.headers.get("Retry-After")
            if retry_after and str(retry_after).isdigit():
                sleep_s = int(retry_after)
            else:
                sleep_s = delay
                delay = min(delay * 2, 60)
            time.sleep(sleep_s)
            continue

        if 500 <= r.status_code <= 599:
            time.sleep(delay)
            delay = min(delay * 2, 60)
            continue

        r.raise_for_status()

    r.raise_for_status()
    return r


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT, "Accept": "application/json"})
    return s


# ============================================================
# 3) Helper: Query dedupe + Query Variants
# ============================================================

def unique_queries(queries: List[str], limit: Optional[int] = None) -> List[str]:
    seen = set()
    out = []
    for q in queries:
        k = re.sub(r"\s+", " ", str(q).strip().lower())
        if k and k not in seen:
            seen.add(k)
            out.append(str(q).strip())
    if limit:
        out = out[:limit]
    return out


_STOPWORDS = {
    "und", "oder", "der", "die", "das", "ein", "eine", "einer", "eines", "mit", "im", "in", "am", "an",
    "the", "and", "or", "of", "for", "in", "on", "to", "a", "an"
}


def tokenize(q: str) -> List[str]:
    q = (q or "").lower()
    toks = re.findall(r"[a-zäöüß0-9]+", q)
    return [t for t in toks if len(t) >= 3 and t not in _STOPWORDS]


def build_query_variants(user_query: str, max_variants: int = 6) -> List[str]:
    """
    Baut robuste Query-Varianten aus einem einzigen User-Query:
    - Original
    - Short (erste 4 Tokens)
    - Phrase (wenn nicht zu lang)
    """
    q = (user_query or "").strip()
    if not q:
        return []

    toks = tokenize(q)
    short = " ".join(toks[:4]) if toks else q

    variants = [q]
    if short and short.lower() != q.lower():
        variants.append(short)

    if 2 <= len(q.split()) <= 6:
        variants.append(f"\"{q}\"")

    # dedupe
    seen = set()
    out = []
    for v in variants:
        k = " ".join(v.lower().split())
        if k and k not in seen:
            seen.add(k)
            out.append(v)
    return out[:max_variants]


# ============================================================
# 4) Text helpers / Screening + Must/Exclude
# ============================================================

NEGATIVE_HINTS = [
    "mind control", "fertility", "sex", "clinical", "cancer", "genome", "protein", "surgery",
    "anatomy", "physiology"
]


def clean_text(text: Optional[str]) -> str:
    if not text:
        return ""
    text = html.unescape(str(text))
    # html-escaped &lt;...&gt; + echte Tags
    text = re.sub(r"&lt;[^&gt;]+&gt;", " ", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_abstract(text: Optional[str]) -> Optional[str]:
    out = clean_text(text)
    return out if out else None


def normalize_doi(doi: Optional[str]) -> Optional[str]:
    if not doi:
        return None
    d = str(doi).strip()
    d = re.sub(r"^https?://(dx\.)?doi\.org/", "", d, flags=re.IGNORECASE).strip()
    return d.lower() if d else None


def passes_min_year(year: Optional[int], min_year: Optional[int]) -> bool:
    if min_year is None:
        return True
    if year is None:
        return True
    return year >= min_year


def passes_max_year(year: Optional[int], max_year: Optional[int]) -> bool:
    if max_year is None:
        return True
    if year is None:
        return True
    return year <= max_year


def _text_blob(*parts: str) -> str:
    return " ".join([p for p in parts if p]).lower()


def contains_any(text: str, terms: List[str]) -> bool:
    if not terms:
        return False
    t = text.lower()
    for x in terms:
        x = str(x).strip().lower()
        if x and x in t:
            return True
    return False


def contains_all(text: str, terms: List[str]) -> bool:
    if not terms:
        return True
    t = text.lower()
    for x in terms:
        x = str(x).strip().lower()
        if x and x not in t:
            return False
    return True


# --- Deine themenspezifischen Screenings (bleiben erhalten für STRICT) ---

def is_text_plausibly_relevant(title: str, abstract: str, keywords: str = "") -> bool:
    """
    Fokus-Screening (dein alter Korridor): Kosten & Projekt ODER Kontext Film/Medien.
    Nur sinnvoll im STRICT-Modus.
    """
    text = f"{title or ''} {abstract or ''} {keywords or ''}".lower()
    if any(bad in text for bad in NEGATIVE_HINTS):
        return False

    cost_hit = any(t in text for t in [
        "kostenrechnung", "teilkostenrechnung", "deckungsbeitrag", "kostenmanagement", "kalkulation",
        "cost accounting", "management accounting", "contribution margin", "variable costing"
    ])
    proj_hit = any(t in text for t in [
        "projektcontrolling", "projektkosten", "projektkostenrechnung", "projektmanagement",
        "project controlling", "project cost", "earned value", "project-based"
    ])
    ctx_hit = any(t in text for t in [
        "filmproduktion", "film production", "creative industries", "medienwirtschaft", "audiovisual"
    ])
    return (cost_hit and proj_hit) or (ctx_hit and (cost_hit or proj_hit))


def openlibrary_is_relevant(title: str, publisher: str, query: str) -> bool:
    """
    Harte OpenLibrary-Filterung (dein alter Korridor).
    Nur sinnvoll im STRICT-Modus.
    """
    text = f"{title or ''} {publisher or ''} {query or ''}".lower()
    if any(bad in text for bad in NEGATIVE_HINTS):
        return False

    cost_hit = any(t in text for t in ["kosten", "cost", "accounting", "deckungsbeitrag", "teilkosten"])
    proj_hit = any(t in text for t in ["projekt", "project", "earned value", "projektcontrolling"])
    ctx_hit = any(t in text for t in ["film", "filmproduktion", "medien", "creative industries"])
    return cost_hit and (proj_hit or ctx_hit)


def dnb_is_relevant_relaxed(title: str, query_original: str) -> bool:
    """
    DNB minimal (dein alter Korridor).
    Im STRICT kann man es nutzen, im NORMAL/OPEN eher nicht.
    """
    text = f"{title or ''} {query_original or ''}".lower()
    if any(bad in text for bad in NEGATIVE_HINTS):
        return False
    return any(t in text for t in [
        "kosten", "cost", "deckungsbeitrag", "teilkosten", "kalkulation",
        "projekt", "project", "projektcontrolling", "project controlling",
        "film", "filmproduktion", "medien", "creative industries"
    ])


# ============================================================
# 5) Record Normalisierung
# ============================================================

def normalize_record(
    title: Optional[str],
    author: Optional[str],
    year: Optional[int],
    doi_or_isbn: Optional[str],
    api_source: str,
    abstract: Optional[str],
    query: str,
    pub_type: Optional[str] = None,
    publisher: Optional[str] = None,
    oa_url: Optional[str] = None,
    url: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "Titel": title,
        "Autor": author,
        "Jahr": year,
        "DOI": doi_or_isbn,
        "URL": url,
        "OA_URL": oa_url,
        "API_Quelle": api_source,
        "Publikationsart": pub_type,
        "Journal_Verlag": publisher,
        "Abstract": abstract,
        "Suchbegriff": query,
        "Keywords": query,
    }


# ============================================================
# 6) Ranking / Scoring
# ============================================================

# --- Dein alter MUST-Gruppen-Score (bleibt für STRICT Profil) ---
MUST_GROUPS = [
    ["kostenrechnung", "teilkostenrechnung", "deckungsbeitrag", "kostenmanagement", "kalkulation",
     "cost accounting", "management accounting", "contribution margin", "variable costing"],
    ["projektcontrolling", "projektkostenrechnung", "projektmanagement", "projektgeschäft",
     "project controlling", "project cost", "earned value", "project-based"],
]
BONUS_GROUP = ["filmproduktion", "film production", "creative industries", "medienwirtschaft", "audiovisual"]


def relevance_score(title: str, abstract: str, keywords: str = "", strict_must: int = 2) -> int:
    text = f"{title or ''} {abstract or ''} {keywords or ''}".lower()
    must_hits = 0
    for group in MUST_GROUPS:
        if any(term in text for term in group):
            must_hits += 1
    if must_hits < strict_must:
        return -999

    score = 0
    for group in MUST_GROUPS:
        score += sum(1 for term in group if term in text)
    score += 3 * sum(1 for term in BONUS_GROUP if term in text)

    if "typology" in text or "typologie" in text:
        score -= 1
    return score


def apply_relevance_filter(df: pd.DataFrame, top_k: int = 500, min_keep: int = 80) -> pd.DataFrame:
    """
    Dein alter Relevanzfilter (MUST-Gruppen). Sinnvoll als STRICT-Profil.
    """
    if df.empty:
        return df

    df = df.copy()
    if "Keywords" not in df.columns:
        df["Keywords"] = df.get("Suchbegriff", "")

    def _run(strict_must: int) -> pd.DataFrame:
        tmp = df.copy()
        tmp["__score"] = tmp.apply(
            lambda r: relevance_score(
                str(r.get("Titel", "")),
                str(r.get("Abstract", "")),
                str(r.get("Keywords", "")),
                strict_must=strict_must
            ),
            axis=1
        )
        return tmp[tmp["__score"] > -100].sort_values("__score", ascending=False)

    out = _run(strict_must=2)
    if len(out) < min_keep:
        out = _run(strict_must=1)

    if out.empty:
        tmp = df.copy()
        tmp["__score"] = tmp.apply(
            lambda r: relevance_score(
                str(r.get("Titel", "")),
                str(r.get("Abstract", "")),
                str(r.get("Keywords", "")),
                strict_must=1
            ),
            axis=1
        )
        tmp = tmp.sort_values("__score", ascending=False)
        out = tmp.head(min_keep)

    out = out.head(top_k)
    out.drop(columns=["__score"], inplace=True, errors="ignore")
    return out


def score_record_query(rec: dict, user_query: str, should_terms: List[str]) -> int:
    """
    Generisches query-basiertes Ranking:
    - Token-Matches im Titel (hoch)
    - Token-Matches im Abstract (mittel)
    - Publisher/Journal (klein)
    - should_terms zusätzliche Boosts
    - Quality: DOI/OA/Autor
    """
    title = (rec.get("Titel") or "").lower()
    abstract = (rec.get("Abstract") or "").lower()
    pub = (rec.get("Journal_Verlag") or "").lower()

    tokens = tokenize(user_query)
    score = 0

    for t in tokens:
        if t in title:
            score += 12
        if t in abstract:
            score += 5
        if t in pub:
            score += 2

    for s in (should_terms or []):
        s = str(s).strip().lower()
        if not s:
            continue
        if s in title:
            score += 8
        elif s in abstract:
            score += 3

    if rec.get("DOI"):
        score += 2
    if rec.get("OA_URL"):
        score += 2
    if rec.get("Autor"):
        score += 1

    return score


def apply_query_ranking_and_filters(
    df: pd.DataFrame,
    *,
    user_query: str,
    mode: str,
    top_k: int,
    must_terms: List[str],
    should_terms: List[str],
    exclude_terms: List[str],
    filters: Dict[str, Any],
) -> pd.DataFrame:
    """
    - Entfernt per exclude_terms
    - Optional: must_terms (in strict/normal)
    - Scoring + Sorting
    - Facettenfilter (year_from/year_to/oa_only/sources/pub_types/publisher_contains)
    """
    if df.empty:
        return df

    mode = (mode or "normal").lower()
    user_query = (user_query or "").strip()

    # Textblob fürs Must/Exclude
    def _blob_row(r: pd.Series) -> str:
        return _text_blob(
            str(r.get("Titel", "") or ""),
            str(r.get("Abstract", "") or ""),
            str(r.get("Journal_Verlag", "") or ""),
            str(r.get("Keywords", "") or ""),
            str(r.get("Suchbegriff", "") or ""),
        )

    df = df.copy()
    df["__blob"] = df.apply(_blob_row, axis=1)

    # Exclude immer
    if exclude_terms:
        df = df[~df["__blob"].apply(lambda t: contains_any(t, exclude_terms))]

    # Must nur in strict/normal (open soll nicht hart filtern)
    if mode in ("strict", "normal") and must_terms:
        df = df[df["__blob"].apply(lambda t: contains_all(t, must_terms))]

    # Facetten: Jahr
    year_from = filters.get("year_from")
    year_to = filters.get("year_to")
    if year_from is not None:
        try:
            yf = int(year_from)
            df = df[(df["Jahr"].isna()) | (df["Jahr"].astype("Int64") >= yf)]
        except Exception:
            pass
    if year_to is not None:
        try:
            yt = int(year_to)
            df = df[(df["Jahr"].isna()) | (df["Jahr"].astype("Int64") <= yt)]
        except Exception:
            pass

    # Facetten: OA only
    if bool(filters.get("oa_only", False)):
        df = df[df.get("OA_URL", "").astype(str).str.strip() != ""]

    # Facetten: Quellen
    sources = filters.get("sources", []) or []
    if sources:
        sset = set([str(s) for s in sources])
        df = df[df.get("API_Quelle", "").astype(str).apply(lambda x: any(s in x for s in sset))]

    # Facetten: Publikationsart
    pub_types = filters.get("pub_types", []) or []
    if pub_types and "Publikationsart" in df.columns:
        pset = set([str(p) for p in pub_types])
        df = df[df["Publikationsart"].astype(str).isin(pset)]

    # Facetten: Publisher/Journal contains
    pub_contains = str(filters.get("publisher_contains", "") or "").strip().lower()
    if pub_contains and "Journal_Verlag" in df.columns:
        df = df[df["Journal_Verlag"].astype(str).str.lower().str.contains(pub_contains, na=False)]

    # Ranking (nur wenn user_query vorhanden)
    if user_query:
        df["__score"] = df.apply(lambda r: score_record_query(r.to_dict(), user_query, should_terms), axis=1)
        df = df.sort_values("__score", ascending=False)

        # Modusabhängige Min-Score Schwelle (optional)
        if mode == "strict":
            min_score = int(filters.get("min_score_strict", 18))
            df = df[df["__score"] >= min_score] if len(df) > 80 else df
        elif mode == "normal":
            min_score = int(filters.get("min_score_normal", 8))
            df = df[df["__score"] >= min_score] if len(df) > 80 else df
        # open: keine Schwelle

        df = df.head(top_k)
        df.drop(columns=["__score"], inplace=True, errors="ignore")
    else:
        # Kein user_query: nur top_k
        df = df.head(top_k)

    df.drop(columns=["__blob"], inplace=True, errors="ignore")
    return df


# ============================================================
# 7) Merge-Best Dedupe (unverändert)
# ============================================================

def _field_score(row: pd.Series) -> int:
    score = 0
    if isinstance(row.get("Abstract"), str) and row["Abstract"].strip():
        score += 8
    if isinstance(row.get("OA_URL"), str) and row["OA_URL"].strip():
        score += 4
    if isinstance(row.get("Autor"), str) and row["Autor"].strip():
        score += 2
    if row.get("Jahr") not in (None, "", 0):
        score += 2
    if isinstance(row.get("Journal_Verlag"), str) and row["Journal_Verlag"].strip():
        score += 1
    if isinstance(row.get("URL"), str) and row["URL"].strip():
        score += 1

    src = str(row.get("API_Quelle", "")).lower()
    if "openalex" in src:
        score += 2
    if "crossref" in src:
        score += 1
    if "semanticscholar" in src:
        score += 1
    return score


def _merge_group(group: pd.DataFrame) -> pd.Series:
    g = group.copy()
    sources = sorted(set([str(x) for x in g.get("API_Quelle", pd.Series(dtype=str)).dropna().astype(str) if x.strip()]))
    sources_str = "; ".join(sources) if sources else None

    g["__score"] = g.apply(_field_score, axis=1)
    best = g.sort_values("__score", ascending=False).iloc[0].copy()

    def pick_first_nonempty(col: str) -> Optional[Any]:
        if col not in g.columns:
            return None
        for v in g[col].tolist():
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return v
        return None

    for col in ["Abstract", "OA_URL", "URL", "Autor", "Journal_Verlag", "Titel", "Jahr", "Publikationsart"]:
        if best.get(col) is None or (isinstance(best.get(col), str) and not best.get(col).strip()):
            cand = pick_first_nonempty(col)
            if cand is not None:
                best[col] = cand

    best["API_Quelle"] = sources_str
    best.drop(labels=["__score"], inplace=True, errors="ignore")
    return best


def groupby_apply_compat(df: pd.DataFrame, by: Union[str, List[str]], func):
    gb = df.groupby(by, group_keys=False)
    try:
        return gb.apply(func, include_groups=False)
    except TypeError:
        return gb.apply(func)


def deduplicate_merge_best(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    df["DOI_norm"] = df.get("DOI", "").astype(str).apply(lambda x: normalize_doi(x) or "").astype(str)
    df["Titel_norm"] = df.get("Titel", "").astype(str).str.strip().str.lower()
    df["Autor_norm"] = df.get("Autor", "").astype(str).str.strip().str.lower()
    df["Jahr_norm"] = df.get("Jahr", "").astype(str).str.strip()

    has_doi = df["DOI_norm"] != ""
    df_doi = df[has_doi].copy()
    df_fb = df[~has_doi].copy()
    df_fb = df_fb[df_fb["Titel_norm"] != ""]

    out_parts = []

    if not df_doi.empty:
        merged = groupby_apply_compat(df_doi, "DOI_norm", _merge_group)
        if isinstance(merged.index, pd.MultiIndex):
            merged = merged.reset_index(drop=True)
        out_parts.append(merged)

    if not df_fb.empty:
        merged_fb = groupby_apply_compat(df_fb, ["Titel_norm", "Jahr_norm", "Autor_norm"], _merge_group)
        if isinstance(merged_fb.index, pd.MultiIndex):
            merged_fb = merged_fb.reset_index(drop=True)
        out_parts.append(merged_fb)

    out = pd.concat(out_parts, ignore_index=True) if out_parts else pd.DataFrame()
    out.drop(columns=["DOI_norm", "Titel_norm", "Autor_norm", "Jahr_norm"], inplace=True, errors="ignore")
    return out


# ============================================================
# 8) API Implementierungen (modi + must/exclude optional)
# ============================================================

def authors_to_string(authors: Optional[List[dict]]) -> Optional[str]:
    if not authors:
        return None
    parts = []
    for a in authors:
        given = a.get("given", "")
        family = a.get("family", "")
        name = ", ".join([p for p in [family, given] if p])
        if name:
            parts.append(name)
    return "; ".join(parts) if parts else None


def year_from_crossref_item(item: dict) -> Optional[int]:
    for field in ("published-print", "published-online", "issued"):
        date_parts = (item.get(field) or {}).get("date-parts")
        if date_parts and isinstance(date_parts, list) and date_parts and isinstance(date_parts[0], list) and date_parts[0]:
            year = date_parts[0][0]
            if isinstance(year, int):
                return year
    return None


def _parse_crossref_query(query_item: Union[str, Dict[str, str]]) -> Tuple[str, Optional[str], Optional[str]]:
    if isinstance(query_item, dict):
        return query_item.get("q", ""), query_item.get("publisher_name"), query_item.get("publisher_location")
    return str(query_item), None, None


def get_crossref_data(
    query_item: Union[str, Dict[str, str]],
    rows: int,
    session: requests.Session,
    min_year: Optional[int],
    max_year: Optional[int] = None,
    *,
    mode: str = "strict",
    must_terms: Optional[List[str]] = None,
    exclude_terms: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    url = "https://api.crossref.org/works"
    q, publisher_name, publisher_location = _parse_crossref_query(query_item)

    params = {
        "query.bibliographic": q,
        "rows": min(max(1, rows), 1000),
        "mailto": MAILTO,
        "select": "DOI,title,author,issued,published-print,published-online,abstract,type,publisher,container-title"
    }
    if publisher_name:
        params["query.publisher-name"] = publisher_name
    if publisher_location:
        params["query.publisher-location"] = publisher_location

    mode = (mode or "strict").lower()
    must_terms = must_terms or []
    exclude_terms = exclude_terms or []

    records: List[Dict[str, Any]] = []
    try:
        r = request_with_backoff(session, "GET", url, api_name="CrossRef", params=params, timeout=HTTP_TIMEOUT)
        items = r.json().get("message", {}).get("items", []) or []

        for item in items:
            title_list = item.get("title", [])
            title = title_list[0] if isinstance(title_list, list) and title_list else None

            year = year_from_crossref_item(item)
            if not passes_min_year(year, min_year) or not passes_max_year(year, max_year):
                continue

            doi = normalize_doi(item.get("DOI"))
            abstract = clean_abstract(item.get("abstract"))
            authors = authors_to_string(item.get("author"))
            pub_type = item.get("type")
            publisher = item.get("publisher")

            blob = _text_blob(title or "", abstract or "", publisher or "", q)

            # Exclude immer
            if contains_any(blob, exclude_terms) or any(bad in blob for bad in NEGATIVE_HINTS):
                continue

            # Must optional in strict/normal
            if mode in ("strict", "normal") and must_terms and not contains_all(blob, must_terms):
                continue

            # Altes Fokus-Screening nur strict
            if mode == "strict":
                if not is_text_plausibly_relevant(title or "", abstract or "", q):
                    continue

            records.append(
                normalize_record(
                    title=title,
                    author=authors,
                    year=year,
                    doi_or_isbn=doi,
                    api_source="CrossRef",
                    abstract=abstract,
                    query=q,
                    pub_type=pub_type,
                    publisher=publisher,
                    url=(f"https://doi.org/{doi}" if doi else None)
                )
            )
        return records

    except Exception:
        return records


def openlibrary_authors_to_string(author_names: Optional[List[str]]) -> Optional[str]:
    if not author_names:
        return None
    return "; ".join([a.strip() for a in author_names if isinstance(a, str) and a.strip()])


def get_openlibrary_data(
    query: str,
    rows: int,
    session: requests.Session,
    min_year: Optional[int],
    max_year: Optional[int] = None,
    *,
    mode: str = "strict",
    must_terms: Optional[List[str]] = None,
    exclude_terms: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    url = "https://openlibrary.org/search.json"
    params = {
        "q": query,
        "limit": rows,
        "lang": "de",
        "fields": "title,author_name,first_publish_year,isbn,publisher"
    }

    mode = (mode or "strict").lower()
    must_terms = must_terms or []
    exclude_terms = exclude_terms or []

    records: List[Dict[str, Any]] = []
    try:
        r = request_with_backoff(session, "GET", url, api_name="OpenLibrary", params=params, timeout=HTTP_TIMEOUT)
        docs = r.json().get("docs", []) or []

        for doc in docs:
            title = doc.get("title") or ""
            year = doc.get("first_publish_year")
            year = year if isinstance(year, int) else None
            if not passes_min_year(year, min_year) or not passes_max_year(year, max_year):
                continue

            publisher_list = doc.get("publisher", []) or []
            publisher = publisher_list[0] if publisher_list else None

            blob = _text_blob(title, publisher or "", query)

            # Exclude immer
            if contains_any(blob, exclude_terms) or any(bad in blob for bad in NEGATIVE_HINTS):
                continue

            # Must optional in strict/normal
            if mode in ("strict", "normal") and must_terms and not contains_all(blob, must_terms):
                continue

            # Harte OpenLibrary-Fokusfilter nur strict
            if mode == "strict":
                if not openlibrary_is_relevant(title, publisher or "", query):
                    continue

            author = openlibrary_authors_to_string(doc.get("author_name"))
            isbn_list = doc.get("isbn", []) or []
            isbn = isbn_list[0] if isbn_list else None

            records.append(
                normalize_record(
                    title=title,
                    author=author,
                    year=year,
                    doi_or_isbn=isbn,
                    api_source="OpenLibrary",
                    abstract=None,
                    query=query,
                    pub_type="Buch",
                    publisher=publisher,
                )
            )
        return records

    except Exception:
        return records


# ---- OpenAlex ----

TRANSLATIONS = {
    "Teilkostenrechnung": ["variable costing"],
    "Deckungsbeitrag": ["contribution margin"],
    "Projektcontrolling": ["project controlling", "project control"],
    "Projektkostenrechnung": ["project cost accounting", "project costing"],
    "Projektgeschäft": ["project-based firm", "project-based organization"],
    "Filmproduktion": ["film production", "audiovisual production", "creative industries"],
    "Kostenrechnung": ["cost accounting", "management accounting"],
    "Kostenmanagement": ["cost management"],
}


def expand_query_bilingual(q: str) -> str:
    parts = [q]
    for de, en_list in TRANSLATIONS.items():
        if de.lower() in q.lower():
            parts.extend(en_list)
    seen = set()
    out = []
    for p in parts:
        k = p.strip().lower()
        if k and k not in seen:
            seen.add(k)
            out.append(p.strip())
    return " ".join(out)


def _reconstruct_openalex_abstract(inv_index: Optional[dict]) -> Optional[str]:
    if not inv_index or not isinstance(inv_index, dict):
        return None
    try:
        max_pos = 0
        for _, positions in inv_index.items():
            if positions:
                max_pos = max(max_pos, max(positions))
        words = [""] * (max_pos + 1)
        for word, positions in inv_index.items():
            for p in positions:
                if 0 <= p < len(words):
                    words[p] = word
        out = " ".join([w for w in words if w])
        return out.strip() or None
    except Exception:
        return None


def get_openalex_data(
    query: str,
    rows: int,
    session: requests.Session,
    min_year: Optional[int],
    max_year: Optional[int] = None,
    *,
    mode: str = "strict",
    must_terms: Optional[List[str]] = None,
    exclude_terms: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    url = "https://api.openalex.org/works"
    qv = expand_query_bilingual(query)

    params = {
        "search": qv,
        "per-page": min(max(1, rows), 200),
        "page": 1,
        "mailto": MAILTO,
        "sort": "cited_by_count:desc",
        "select": "id,display_name,publication_year,doi,type,authorships,primary_location,abstract_inverted_index"
    }
    if OPENALEX_API_KEY:
        params["api_key"] = OPENALEX_API_KEY

    mode = (mode or "strict").lower()
    must_terms = must_terms or []
    exclude_terms = exclude_terms or []

    records: List[Dict[str, Any]] = []
    try:
        r = request_with_backoff(session, "GET", url, api_name="OpenAlex", params=params, timeout=HTTP_TIMEOUT)
        results = r.json().get("results", []) or []

        for w in results:
            title = (w.get("display_name") or "").strip()
            year = w.get("publication_year")
            year = year if isinstance(year, int) else None
            if not passes_min_year(year, min_year) or not passes_max_year(year, max_year):
                continue

            doi = normalize_doi(w.get("doi"))

            primary_loc = w.get("primary_location") or {}
            source = primary_loc.get("source") or {}
            publisher = source.get("display_name")

            abstract = _reconstruct_openalex_abstract(w.get("abstract_inverted_index")) or ""
            pub_type = w.get("type")

            author_names = []
            for a in (w.get("authorships") or []):
                n = ((a.get("author") or {}).get("display_name") or "").strip()
                if n:
                    author_names.append(n)
            author = "; ".join(author_names) if author_names else None

            blob = _text_blob(title, abstract, publisher or "", qv)

            if contains_any(blob, exclude_terms) or any(bad in blob for bad in NEGATIVE_HINTS):
                continue

            if mode in ("strict", "normal") and must_terms and not contains_all(blob, must_terms):
                continue

            if mode == "strict":
                if not is_text_plausibly_relevant(title, abstract, qv):
                    continue

            records.append(
                normalize_record(
                    title=title,
                    author=author,
                    year=year,
                    doi_or_isbn=doi,
                    api_source="OpenAlex",
                    abstract=(abstract if abstract else None),
                    query=qv,
                    pub_type=pub_type,
                    publisher=publisher,
                    url=(f"https://doi.org/{doi}" if doi else None)
                )
            )
        return records

    except Exception:
        return records


# ---- DNB SRU ----

def build_dnb_cql(q: str, max_terms: int = 4) -> str:
    q = q.strip()
    if not q:
        return q
    tokens = [t for t in re.split(r"\s+", q) if len(t.strip()) >= 3]
    tokens = tokens[:max_terms] if tokens else [q]
    return " and ".join([f'WOE="{t}"' for t in tokens])


def get_dnb_sru_data(
    query: str,
    rows: int,
    session: requests.Session,
    min_year: Optional[int],
    max_year: Optional[int] = None,
    *,
    mode: str = "strict",
    must_terms: Optional[List[str]] = None,
    exclude_terms: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    base_url = "https://services.dnb.de/sru/dnb"
    cql = build_dnb_cql(query, max_terms=4)

    params = {
        "version": "1.1",
        "operation": "searchRetrieve",
        "query": cql,
        "maximumRecords": str(min(max(1, rows), 100)),
        "startRecord": "1",
        "recordSchema": "oai_dc",
    }

    mode = (mode or "strict").lower()
    must_terms = must_terms or []
    exclude_terms = exclude_terms or []

    records: List[Dict[str, Any]] = []
    try:
        r = request_with_backoff(session, "GET", base_url, api_name="DNB_SRU", params=params, timeout=HTTP_TIMEOUT)
        xml_text = r.text

        ns = {"srw": "http://www.loc.gov/zing/srw/",
              "dc": "http://purl.org/dc/elements/1.1/",
              "oai_dc": "http://www.openarchives.org/OAI/2.0/oai_dc/"}

        root = ET.fromstring(xml_text)
        for rec in root.findall(".//srw:record", ns):
            record_data = rec.find(".//srw:recordData", ns)
            if record_data is None:
                continue
            dc_node = record_data.find(".//oai_dc:dc", ns)
            if dc_node is None:
                continue

            title_node = dc_node.find("dc:title", ns)
            title = title_node.text.strip() if title_node is not None and title_node.text else None

            creators = [c.text.strip() for c in dc_node.findall("dc:creator", ns) if c is not None and c.text]
            author = "; ".join(creators) if creators else None

            year = None
            d = dc_node.find("dc:date", ns)
            if d is not None and d.text:
                m = re.search(r"(\d{4})", d.text)
                if m:
                    year = int(m.group(1))
            if not passes_min_year(year, min_year) or not passes_max_year(year, max_year):
                continue

            identifiers = [i.text.strip() for i in dc_node.findall("dc:identifier", ns) if i is not None and i.text]
            isbn = None
            for ident in identifiers:
                if "isbn" in ident.lower():
                    isbn = ident.split(":")[-1].strip()
                    break

            publisher_node = dc_node.find("dc:publisher", ns)
            publisher = publisher_node.text.strip() if publisher_node is not None and publisher_node.text else None

            blob = _text_blob(title or "", publisher or "", query)

            if contains_any(blob, exclude_terms) or any(bad in blob for bad in NEGATIVE_HINTS):
                continue

            if mode in ("strict", "normal") and must_terms and not contains_all(blob, must_terms):
                continue

            # DNB Fokusfilter nur strict
            if mode == "strict":
                if not dnb_is_relevant_relaxed(title or "", query):
                    continue

            records.append(
                normalize_record(
                    title=title,
                    author=author,
                    year=year,
                    doi_or_isbn=isbn,
                    api_source="DNB_SRU",
                    abstract=None,
                    query=query,
                    pub_type="Buch",
                    publisher=publisher,
                    url=None
                )
            )
        return records

    except Exception:
        return records


# ---- Unpaywall ----

def enrich_with_unpaywall(df: pd.DataFrame, session: requests.Session) -> pd.DataFrame:
    if df.empty:
        return df
    if not UNPAYWALL_EMAIL:
        return df

    df = df.copy()
    if "OA_URL" not in df.columns:
        df["OA_URL"] = None

    dois = sorted(set(
        normalize_doi(d) for d in df.get("DOI", pd.Series(dtype=str)).dropna().astype(str).tolist()
        if normalize_doi(d)
    ))
    if not dois:
        return df

    cache: Dict[str, Optional[str]] = {}
    for doi in dois:
        url = f"https://api.unpaywall.org/v2/{doi}"
        params = {"email": UNPAYWALL_EMAIL}
        try:
            r = request_with_backoff(session, "GET", url, api_name="Unpaywall", params=params, timeout=HTTP_TIMEOUT)
            if r.status_code == 404:
                cache[doi] = None
                continue
            data = r.json()
            best = data.get("best_oa_location")
            oa_url = None
            if isinstance(best, dict):
                oa_url = best.get("url_for_pdf") or best.get("url") or best.get("url_for_landing_page")
            cache[doi] = oa_url if isinstance(oa_url, str) and oa_url.strip() else None
        except Exception:
            cache[doi] = None

    df["DOI_norm_tmp"] = df["DOI"].astype(str).apply(lambda x: normalize_doi(x))
    df["OA_URL"] = df.apply(
        lambda r: r["OA_URL"] if (isinstance(r.get("OA_URL"), str) and r.get("OA_URL").strip())
        else cache.get(r.get("DOI_norm_tmp")),
        axis=1
    )
    df.drop(columns=["DOI_norm_tmp"], inplace=True, errors="ignore")
    return df


# ---- Semantic Scholar DOI Enrichment (Cache als Dict) ----

def _semscholar_headers() -> Dict[str, str]:
    h = {"Accept": "application/json", "User-Agent": USER_AGENT}
    if SEMANTIC_SCHOLAR_API_KEY:
        h["x-api-key"] = SEMANTIC_SCHOLAR_API_KEY
    return h


def semscholar_enrich_by_doi(
    dois: List[str],
    session: requests.Session,
    *,
    s2_cache: Optional[Dict[str, dict]] = None,
    min_year: Optional[int] = None,
    max_year: Optional[int] = None,
    mode: str = "strict",
    must_terms: Optional[List[str]] = None,
    exclude_terms: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    DOI-Enrichment: holt Abstract/OA PDF etc. pro DOI.
    Filterung:
    - exclude always
    - must in strict/normal
    - Fokus-Screening nur strict
    """
    out: List[Dict[str, Any]] = []
    cache = s2_cache if isinstance(s2_cache, dict) else {}

    mode = (mode or "strict").lower()
    must_terms = must_terms or []
    exclude_terms = exclude_terms or []

    for doi in dois:
        try:
            if doi in cache:
                p = cache[doi]
            else:
                url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}"
                params = {"fields": "title,authors,year,venue,abstract,url,externalIds,openAccessPdf"}
                r = request_with_backoff(
                    session, "GET", url,
                    api_name="SemanticScholar",
                    params=params,
                    headers=_semscholar_headers(),
                    timeout=HTTP_TIMEOUT
                )
                p = r.json()
                cache[doi] = p

            title = p.get("title")
            year = p.get("year")
            year = year if isinstance(year, int) else None
            if not passes_min_year(year, min_year) or not passes_max_year(year, max_year):
                continue

            authors = p.get("authors") or []
            author = "; ".join([a.get("name") for a in authors if a.get("name")]) if authors else None

            venue = p.get("venue")
            abstract = clean_abstract(p.get("abstract"))
            paper_url = p.get("url")
            oa_url = (p.get("openAccessPdf") or {}).get("url")
            doi2 = normalize_doi(((p.get("externalIds") or {}).get("DOI")) or doi)

            blob = _text_blob(title or "", abstract or "", venue or "", doi)

            if contains_any(blob, exclude_terms) or any(bad in blob for bad in NEGATIVE_HINTS):
                continue

            if mode in ("strict", "normal") and must_terms and not contains_all(blob, must_terms):
                continue

            if mode == "strict":
                if not is_text_plausibly_relevant(title or "", abstract or "", "doi_enrich"):
                    continue

            out.append(
                normalize_record(
                    title=title,
                    author=author,
                    year=year,
                    doi_or_isbn=doi2,
                    api_source="SemanticScholar",
                    abstract=abstract,
                    query=f"DOI_ENRICH:{doi}",
                    pub_type="Paper",
                    publisher=venue,
                    oa_url=(oa_url if isinstance(oa_url, str) else None),
                    url=paper_url
                )
            )
        except Exception:
            continue

    return out


# ============================================================
# 9) Excel: In-Memory Writer (unverändert)
# ============================================================

def _norm_header(s: str) -> str:
    return " ".join(str(s).strip().lower().replace("\n", " ").split())


def _build_header_map(ws, header_row: int = 1) -> dict:
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        key = _norm_header(val)
        if key and key not in header_map:
            header_map[key] = col
    return header_map


def _find_first_free_row(ws, title_col: int, start_row: int = 2) -> int:
    r = start_row
    while r <= ws.max_row:
        v = ws.cell(row=r, column=title_col).value
        if v is None or str(v).strip() == "" or str(v).strip() == "0":
            return r
        r += 1
    return ws.max_row + 1


def _max_nr(ws, nr_col: int, start_row: int = 2) -> int:
    m = 0
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=nr_col).value
        try:
            m = max(m, int(v))
        except Exception:
            pass
    return m


def write_to_excel_bytes(df: pd.DataFrame, template_bytes: bytes, sheet_name: str, header_row: int = 1) -> bytes:
    if df.empty:
        return template_bytes

    bio_in = BytesIO(template_bytes)
    wb = load_workbook(bio_in)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' existiert nicht. Vorhanden: {wb.sheetnames}")

    ws = wb[sheet_name]
    header_map = _build_header_map(ws, header_row=header_row)

    def col_of(*variants):
        for v in variants:
            key = _norm_header(v)
            if key in header_map:
                return header_map[key]
        return None

    col_nr = col_of("Nr")
    col_title = col_of("Titel")
    col_author = col_of("Autor(en)", "Autor")
    col_year = col_of("Erscheinungsjahr", "Erscheinungjahr")
    col_doi = col_of("DOI / URL", "DOI", "URL")
    col_pubtype = col_of("Publikationsart")
    col_kw = col_of("Keywords")
    col_abs = col_of("Abstract")
    col_journal = col_of("Journal / Verlag", "Journal/Verlag", "Verlag")
    col_api = col_of("API Quelle", "API-Quelle", "Quelle")
    col_oa = col_of("OA URL", "Open Access URL", "Volltext URL", "Fulltext URL", "OA-Link", "OpenAccess")

    if col_title is None:
        raise ValueError("Konnte die Spalte 'Titel' im Template nicht finden.")

    next_nr = _max_nr(ws, col_nr, start_row=header_row + 1) + 1 if col_nr else None
    current_row = _find_first_free_row(ws, title_col=col_title, start_row=header_row + 1)

    needed = ["Titel", "Autor", "Jahr", "DOI", "URL", "OA_URL", "Publikationsart",
              "Keywords", "Abstract", "Journal_Verlag", "API_Quelle"]
    for c in needed:
        if c not in df.columns:
            df[c] = None
    dfw = df[needed].copy()

    for r in dfw.itertuples(index=False):
        target_row = current_row

        if col_nr and (ws.cell(row=target_row, column=col_nr).value in [None, "", "0", 0]):
            ws.cell(row=target_row, column=col_nr).value = next_nr
            next_nr += 1

        ws.cell(row=target_row, column=col_title).value = getattr(r, "Titel", None)

        if col_author:
            ws.cell(row=target_row, column=col_author).value = getattr(r, "Autor", None)
        if col_year:
            ws.cell(row=target_row, column=col_year).value = getattr(r, "Jahr", None)

        if col_doi:
            doi_val = getattr(r, "DOI", None)
            url_val = getattr(r, "URL", None)
            oa_val = getattr(r, "OA_URL", None)
            cell_val = doi_val or url_val
            if (col_oa is None) and oa_val and isinstance(oa_val, str) and oa_val.strip():
                if cell_val and isinstance(cell_val, str) and cell_val.strip():
                    cell_val = f"{cell_val}\nOA: {oa_val}"
                else:
                    cell_val = oa_val
            ws.cell(row=target_row, column=col_doi).value = cell_val

        if col_oa:
            ws.cell(row=target_row, column=col_oa).value = getattr(r, "OA_URL", None)

        if col_pubtype:
            ws.cell(row=target_row, column=col_pubtype).value = getattr(r, "Publikationsart", None)
        if col_api:
            ws.cell(row=target_row, column=col_api).value = getattr(r, "API_Quelle", None)
        if col_journal:
            ws.cell(row=target_row, column=col_journal).value = getattr(r, "Journal_Verlag", None)
        if col_abs:
            ws.cell(row=target_row, column=col_abs).value = getattr(r, "Abstract", None)
        if col_kw:
            ws.cell(row=target_row, column=col_kw).value = getattr(r, "Keywords", None)

        current_row += 1

    bio_out = BytesIO()
    wb.save(bio_out)
    return bio_out.getvalue()


# ============================================================
# 10) Hauptworkflow: run_search_to_df (erweitert)
# ============================================================

def run_search_to_df(
    config: Dict[str, Any],
    *,
    s2_cache: Optional[Dict[str, dict]] = None,
    progress_cb=None
) -> pd.DataFrame:
    """
    Backward compatible:
    - Wenn du nur 'queries' gibst, läuft es wie vorher (STRICT default).
    Neu:
    - mode: 'strict'|'normal'|'open'
    - user_query + must/should/exclude
    - filters: Facettenfilter
    """
    session = make_session()
    all_records: List[Dict[str, Any]] = []

    # Basiseinstellungen
    min_year = int(config.get("min_year", 2000))
    top_k = int(config.get("top_k", 500))

    rows_crossref = int(config.get("rows_crossref", 200))
    rows_openlib = int(config.get("rows_openlib", 50))
    rows_openalex = int(config.get("rows_openalex", 120))
    rows_dnb = int(config.get("rows_dnb", 15))

    enable_crossref = bool(config.get("enable_crossref", True))
    enable_openlib = bool(config.get("enable_openlib", True))
    enable_openalex = bool(config.get("enable_openalex", True))
    enable_dnb = bool(config.get("enable_dnb", True))

    use_publisher_hints = bool(config.get("use_publisher_hints", True))

    enable_unpaywall = bool(config.get("enable_unpaywall", True))
    enable_s2 = bool(config.get("enable_s2", False))
    s2_topn = int(config.get("s2_topn", 60))
    s2_only_if_missing = bool(config.get("s2_only_if_missing", True))

    # Neue Steuerung
    mode = str(config.get("mode", "strict")).lower().strip()  # strict|normal|open
    if mode not in ("strict", "normal", "open"):
        mode = "strict"

    user_query = str(config.get("user_query", "") or "").strip()
    must_terms = [str(x).strip() for x in (config.get("must_terms", []) or []) if str(x).strip()]
    should_terms = [str(x).strip() for x in (config.get("should_terms", []) or []) if str(x).strip()]
    exclude_terms = [str(x).strip() for x in (config.get("exclude_terms", []) or []) if str(x).strip()]

    filters = config.get("filters", {}) or {}
    max_year = filters.get("year_to")  # optional
    # year_from (=min_year) bleibt min_year gesteuert

    # Queries (klassisch) oder automatisch aus user_query
    q = config.get("queries", {}) or {}
    q_crossref = unique_queries(q.get("crossref", []) or [])
    q_openlib = unique_queries(q.get("openlib", []) or [])
    q_openalex = unique_queries(q.get("openalex", []) or [])
    q_dnb = unique_queries(q.get("dnb", []) or [])

    # Wenn keine Queries übergeben wurden, aber user_query existiert:
    if user_query and (len(q_crossref) + len(q_openlib) + len(q_openalex) + len(q_dnb) == 0):
        variants = build_query_variants(user_query, max_variants=6)
        q_crossref = variants[:] if enable_crossref else []
        q_openalex = variants[:] if enable_openalex else []
        q_openlib = variants[:] if enable_openlib else []
        q_dnb = variants[:] if enable_dnb else []

    # Wenn user_query leer ist, aber Queries existieren: nimm erste Query als Ranking-Query
    if not user_query:
        for lst in (q_crossref, q_openalex, q_openlib, q_dnb):
            if lst:
                user_query = lst[0]
                break

    def say(m: str):
        if callable(progress_cb):
            progress_cb(m)

    # --- CrossRef ---
    if enable_crossref and q_crossref:
        for query in q_crossref:
            q_item: Union[str, Dict[str, str]] = query
            if use_publisher_hints and ("Projektcontrolling" in query or "Kostenrechnung" in query):
                q_item = {"q": query, "publisher_name": PUBLISHER_NAME_HINTS[0]}
            say(f"CrossRef: {query}")
            recs = get_crossref_data(
                q_item,
                rows=rows_crossref,
                session=session,
                min_year=min_year,
                max_year=max_year,
                mode=mode,
                must_terms=must_terms,
                exclude_terms=exclude_terms
            )
            all_records.extend(recs)

    # --- OpenLibrary ---
    if enable_openlib and q_openlib:
        for query in q_openlib:
            say(f"OpenLibrary: {query}")
            recs = get_openlibrary_data(
                query,
                rows=rows_openlib,
                session=session,
                min_year=min_year,
                max_year=max_year,
                mode=mode,
                must_terms=must_terms,
                exclude_terms=exclude_terms
            )
            all_records.extend(recs)

    # --- OpenAlex ---
    if enable_openalex and q_openalex:
        for query in q_openalex:
            say(f"OpenAlex: {query}")
            recs = get_openalex_data(
                query,
                rows=rows_openalex,
                session=session,
                min_year=min_year,
                max_year=max_year,
                mode=mode,
                must_terms=must_terms,
                exclude_terms=exclude_terms
            )
            all_records.extend(recs)

    # --- DNB SRU ---
    if enable_dnb and q_dnb:
        for query in q_dnb:
            say(f"DNB SRU: {query}")
            recs = get_dnb_sru_data(
                query,
                rows=rows_dnb,
                session=session,
                min_year=min_year,
                max_year=max_year,
                mode=mode,
                must_terms=must_terms,
                exclude_terms=exclude_terms
            )
            all_records.extend(recs)

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame(all_records)

    # 1) Merge-Dedupe
    df = deduplicate_merge_best(df)

    # 2) Ranking/Filter abhängig vom Modus
    if mode == "strict":
        # Dein bisheriger Fokus-Filter + TopK
        df = apply_relevance_filter(df, top_k=top_k, min_keep=80)
    else:
        # Generisches query-basiertes Ranking + Facettenfilter
        df = apply_query_ranking_and_filters(
            df,
            user_query=user_query,
            mode=mode,
            top_k=top_k,
            must_terms=must_terms,
            should_terms=should_terms,
            exclude_terms=exclude_terms,
            filters=filters
        )

    # 3) Semantic Scholar DOI Enrichment (optional)
    if enable_s2 and not df.empty:
        dois = []
        if "DOI" in df.columns:
            for _, r in df.iterrows():
                doi = normalize_doi(r.get("DOI"))
                if not doi:
                    continue

                if s2_only_if_missing:
                    has_abs = isinstance(r.get("Abstract"), str) and r.get("Abstract").strip()
                    has_oa = isinstance(r.get("OA_URL"), str) and r.get("OA_URL").strip()
                    if has_abs and has_oa:
                        continue

                if doi not in dois:
                    dois.append(doi)
                if len(dois) >= s2_topn:
                    break

        if dois:
            say(f"SemanticScholar Enrichment: {len(dois)} DOIs")
            extra = semscholar_enrich_by_doi(
                dois,
                session=session,
                s2_cache=s2_cache,
                min_year=min_year,
                max_year=max_year,
                mode=mode,
                must_terms=must_terms,
                exclude_terms=exclude_terms
            )
            if extra:
                df = pd.concat([df, pd.DataFrame(extra)], ignore_index=True)
                df = deduplicate_merge_best(df)

                # Nach-Enrichment erneut ranken (nur in normal/open)
                if mode != "strict":
                    df = apply_query_ranking_and_filters(
                        df,
                        user_query=user_query,
                        mode=mode,
                        top_k=top_k,
                        must_terms=must_terms,
                        should_terms=should_terms,
                        exclude_terms=exclude_terms,
                        filters=filters
                    )
                else:
                    df = apply_relevance_filter(df, top_k=top_k, min_keep=80)

    # 4) Unpaywall
    if enable_unpaywall and not df.empty:
        df = enrich_with_unpaywall(df, session=session)

    # 5) final merge-dedupe
    df = deduplicate_merge_best(df)

    # 6) final top_k safeguard
    if len(df) > top_k:
        df = df.head(top_k)

    return df
