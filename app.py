import hashlib
import json
from io import BytesIO
import numpy as np

import streamlit as st
from openpyxl import load_workbook, Workbook

from core_literatur_recherche9 import run_search_to_df, write_to_excel_bytes


# ============================================================
# WICHTIG (Deployment-Hinweis):
# Um das Streamlit-Menü inkl. "View source" zuverlässig zu entfernen,
# lege zusätzlich in deinem Repo diese Datei an:
#
#   .streamlit/config.toml
#
# mit folgendem Inhalt:
#   [client]
#   toolbarMode = "minimal"
#
# (Dann zeigt Streamlit nur noch extern/über set_page_config definierte
# Menüeinträge – und wenn keine da sind, wird das Menü ausgeblendet.)
# ============================================================


# ----------------------------
# Page setup
# ----------------------------
st.set_page_config(
    page_title="LiteraturrechercheTool",
    layout="wide",
    # In Kombination mit client.toolbarMode="minimal" sorgt ein leeres menu_items
    # dafür, dass kein Menü (und damit auch kein "View source") angezeigt wird.
    menu_items={}
)

# UI-Hardening (nicht 100% gegen DevTools, aber entfernt die "bequemen" Wege)
st.markdown(
    """
    <style>
      /* Streamlit "Chrome" ausblenden (Toolbar/Header/Footer) */
      div[data-testid="stToolbar"] {visibility: hidden !important; height: 0px !important; position: fixed !important;}
      div[data-testid="stHeader"] {visibility: hidden !important; height: 0px !important; position: fixed !important;}
      #MainMenu {visibility: hidden !important;}
      footer {visibility: hidden !important;}

      /* Copy-to-clipboard Buttons in Code-Blöcken ausblenden (falls vorhanden) */
      button[data-testid="stCopyToClipboardButton"] {display: none !important;}

      /* Optional: obere Deko-Leiste (variiert je nach Streamlit-Version) */
      div[data-testid="stDecoration"] {visibility: hidden !important; height: 0px !important;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("LiteraturrechercheTool")


# ----------------------------
# Session state init
# ----------------------------
if "s2_cache" not in st.session_state:
    st.session_state["s2_cache"] = {}

if "last_config_hash" not in st.session_state:
    st.session_state["last_config_hash"] = None

if "last_df" not in st.session_state:
    st.session_state["last_df"] = None

if "last_sheet" not in st.session_state:
    st.session_state["last_sheet"] = None

if "last_template_bytes" not in st.session_state:
    st.session_state["last_template_bytes"] = None


# ----------------------------
# Option 3: Template mit Header erzeugen (Download)
# ----------------------------
DEFAULT_TEMPLATE_HEADERS = [
    "Nr",
    "Titel",
    "Autor(en)",
    "Erscheinungsjahr",
    "DOI / URL",
    "Publikationsart",
    "Keywords",
    "Abstract",
    "Journal / Verlag",
    "API Quelle",
    "OA URL",
]


def make_template_bytes(sheet_name: str = "Final_Thema2", header_row: int = 1) -> bytes:
    """Erstellt ein neues Excel-Template mit einem Sheet und Header-Zeile."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for _ in range(1, int(header_row)):
        ws.append([])

    ws.append(DEFAULT_TEMPLATE_HEADERS)

    widths = {
        "A": 6,   # Nr
        "B": 45,  # Titel
        "C": 28,  # Autor(en)
        "D": 14,  # Jahr
        "E": 28,  # DOI/URL
        "F": 16,  # Publikationsart
        "G": 30,  # Keywords
        "H": 45,  # Abstract
        "I": 26,  # Journal/Verlag
        "J": 18,  # API Quelle
        "K": 30,  # OA URL
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


st.subheader("0) Optional: Leeres Template herunterladen (empfohlen)")
with st.expander("📥 Excel-Template mit korrekten Headern erstellen", expanded=False):
    st.write(
        "Wenn dein eigenes Template keine Spalte **„Titel“** hat oder die Header in einer anderen Zeile stehen, "
        "kannst du hier ein frisches Template herunterladen, das garantiert kompatibel ist."
    )
    t1, t2, t3 = st.columns([2, 1, 1])
    with t1:
        template_sheet_name = st.text_input("Sheet-Name im Template", value="Final_Thema2")
    with t2:
        template_header_row = st.number_input("Header-Zeile im Template", min_value=1, max_value=50, value=1, step=1)
    with t3:
        st.write("")
    st.download_button(
        "Template herunterladen (.xlsx)",
        data=make_template_bytes(sheet_name=template_sheet_name, header_row=int(template_header_row)),
        file_name="literatur_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------
# Defaults (Queries) – Expertenmodus
# ----------------------------
DEFAULT_CROSSREF = [
    "Teilkostenrechnung Projektcontrolling",
    "Deckungsbeitrag Projektcontrolling",
    "Projektkostenrechnung Teilkostenrechnung",
    "Projektkostenrechnung Deckungsbeitrag",
    "Projektcontrolling Earned Value",
    "Projektcontrolling Budgetierung",
    "Projektkostenrechnung Kalkulation",
    "Filmproduktion Projektcontrolling",
    "Filmproduktion Kostenmanagement",
    "Creative Industries project controlling",
    "project-based firm management accounting",
    "Projektcontrolling Kontingenzansatz",
    "Branchenvergleich Projektcontrolling",
]

DEFAULT_OPENLIB = [
    "Projektcontrolling Kostenrechnung",
    "Projektkostenrechnung Kostenmanagement",
    "Teilkostenrechnung Deckungsbeitrag",
    "Kosten- und Leistungsrechnung Projekt",
    "Projektmanagement Kosten Controlling",
    "Projektcontrolling Earned Value",
    "Kostenmanagement im Projekt",
    "Projektgeschäft Controlling",
    "Filmproduktion Kostenmanagement",
    "Filmproduktion Projektcontrolling",
    "Medienwirtschaft Controlling Kosten",
    "Creative industries project controlling",
    "Projektcontrolling Deutschland",
    "Projektcontrolling Österreich",
]

DEFAULT_OPENALEX = [
    "Projektcontrolling Kostenrechnung",
    "Teilkostenrechnung Projektcontrolling",
    "Deckungsbeitrag Projektcontrolling",
    "Kostenmanagement Projektcontrolling",
    "Projektkostenrechnung",
    "earned value project controlling",
    "project-based firm management accounting",
    "creative industries project controlling",
    "film production project controlling",
]

DEFAULT_DNB = [
    "Projektcontrolling Kostenrechnung",
    "Projektkostenrechnung",
    "Kostenmanagement Projekt",
    "Kosten- und Leistungsrechnung",
    "Teilkostenrechnung",
    "Deckungsbeitrag",
    "Projektgeschäft Controlling",
    "Filmproduktion Controlling",
    "Medienwirtschaft Controlling",
]


def _dedupe_lines(text: str) -> list[str]:
    """1 Query pro Zeile, trim + dedupe (case/whitespace-insensitive)."""
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    seen = set()
    out = []
    for q in lines:
        k = " ".join(q.lower().split())
        if k not in seen:
            seen.add(k)
            out.append(q)
    return out


def _split_csv(s: str) -> list[str]:
    return [x.strip() for x in (s or "").split(",") if x.strip()]


def _ensure_state():
    # Query textareas (Expert mode)
    if "q_text_crossref" not in st.session_state:
        st.session_state["q_text_crossref"] = "\n".join(DEFAULT_CROSSREF)
    if "q_text_openlib" not in st.session_state:
        st.session_state["q_text_openlib"] = "\n".join(DEFAULT_OPENLIB)
    if "q_text_openalex" not in st.session_state:
        st.session_state["q_text_openalex"] = "\n".join(DEFAULT_OPENALEX)
    if "q_text_dnb" not in st.session_state:
        st.session_state["q_text_dnb"] = "\n".join(DEFAULT_DNB)

    # Generic search controls
    if "user_query" not in st.session_state:
        st.session_state["user_query"] = ""
    if "mode_label" not in st.session_state:
        st.session_state["mode_label"] = "Normal"
    if "must_text" not in st.session_state:
        st.session_state["must_text"] = ""
    if "should_text" not in st.session_state:
        st.session_state["should_text"] = ""
    if "exclude_text" not in st.session_state:
        st.session_state["exclude_text"] = "medizin, fertility, cancer, genome"  # hilfreicher Start

    # Facets
    if "year_to" not in st.session_state:
        st.session_state["year_to"] = 2100
    if "oa_only" not in st.session_state:
        st.session_state["oa_only"] = False
    if "sources_filter" not in st.session_state:
        st.session_state["sources_filter"] = []
    if "pub_types_filter" not in st.session_state:
        st.session_state["pub_types_filter"] = []
    if "publisher_contains" not in st.session_state:
        st.session_state["publisher_contains"] = ""


def _reset_queries_to_defaults():
    st.session_state["q_text_crossref"] = "\n".join(DEFAULT_CROSSREF)
    st.session_state["q_text_openlib"] = "\n".join(DEFAULT_OPENLIB)
    st.session_state["q_text_openalex"] = "\n".join(DEFAULT_OPENALEX)
    st.session_state["q_text_dnb"] = "\n".join(DEFAULT_DNB)


_ensure_state()


# ----------------------------
# UI: Upload
# ----------------------------
st.subheader("1) Excel Template hochladen")
uploaded = st.file_uploader("Excel (.xlsx)", type=["xlsx"])

if not uploaded:
    st.info("Bitte eine Excel-Datei hochladen (oder oben ein Template herunterladen).")
    st.stop()

template_bytes = uploaded.getvalue()

try:
    wb = load_workbook(BytesIO(template_bytes), read_only=False, data_only=False)
except Exception as e:
    st.error(f"Excel konnte nicht gelesen werden: {e}")
    st.stop()

sheet_name = st.selectbox("Sheet auswählen", wb.sheetnames, index=0)
header_row = st.number_input("Header-Zeile (meist 1)", min_value=1, max_value=50, value=1, step=1)

with st.expander("🔎 Debug: Header in der gewählten Header-Zeile anzeigen", expanded=False):
    ws = wb[sheet_name]
    row = int(header_row)
    headers = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
    st.write(headers)
    st.info("Wenn 'Titel' nicht exakt vorkommt, nutze das Template oben oder ändere Header-Zeile/Sheet.")


# ----------------------------
# UI: Einstellungen in einem Form (damit Run nur beim Submit)
# ----------------------------
st.subheader("2) Suche & Parameter (Generisch + Präzise Filter)")

with st.form("settings_form", clear_on_submit=False):

    # --- Generic search controls ---
    st.markdown("### A) Suchsteuerung (empfohlen)")
    g1, g2 = st.columns([2, 1])
    with g1:
        user_query = st.text_input(
            "Hauptsuchbegriff (z.B. Steuerberatung, ESG Reporting, Supply Chain ...)",
            value=st.session_state["user_query"],
            placeholder="z.B. Steuerberatung",
        )
    with g2:
        mode_label = st.selectbox(
            "Suchmodus",
            ["Offen", "Normal", "Streng"],
            index=["Offen", "Normal", "Streng"].index(st.session_state["mode_label"]),
            help="Offen: mehr Treffer, Normal: ausgewogen, Streng: Fokusfilter + härtere Schwellen",
        )

    f1, f2, f3 = st.columns(3)
    with f1:
        must_text = st.text_input(
            "MUSS enthalten (Komma-separiert, optional)",
            value=st.session_state["must_text"],
            placeholder="z.B. steuerberatung, tax advisory",
        )
    with f2:
        should_text = st.text_input(
            "SOLL enthalten (Komma-separiert, optional)",
            value=st.session_state["should_text"],
            placeholder="z.B. digitalisierung, compliance",
        )
    with f3:
        exclude_text = st.text_input(
            "AUSSCHLIESSEN (Komma-separiert)",
            value=st.session_state["exclude_text"],
            placeholder="z.B. medizin, biology, fertility",
        )

    st.markdown("### B) Quellen & Limits")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        enable_crossref = st.checkbox("CrossRef", value=True)
    with c2:
        enable_openalex = st.checkbox("OpenAlex", value=True)
    with c3:
        enable_openlib = st.checkbox("OpenLibrary", value=True)
    with c4:
        enable_dnb = st.checkbox("DNB SRU", value=True)

    p1, p2, p3, p4 = st.columns(4)
    with p1:
        min_year = st.number_input("Min. Jahr", 1900, 2100, 2000)
    with p2:
        top_k = st.number_input("Top-K (final)", 50, 5000, 500)
    with p3:
        rows_crossref = st.slider("Rows CrossRef", 10, 1000, 200)
    with p4:
        rows_openalex = st.slider("Rows OpenAlex", 10, 200, 120)

    p5, p6, p7, p8 = st.columns(4)
    with p5:
        rows_openlib = st.slider("Rows OpenLibrary", 5, 200, 50)
    with p6:
        rows_dnb = st.slider("Rows DNB", 5, 100, 15)
    with p7:
        use_publisher_hints = st.checkbox("CrossRef Publisher-Hints", value=True)
    with p8:
        enable_unpaywall = st.checkbox("Unpaywall OA_URL", value=True)

    st.markdown("### C) Präzise Facettenfilter (wirken nach dem Abruf/Ranking)")
    ff1, ff2, ff3, ff4 = st.columns(4)
    with ff1:
        year_to = st.number_input("Jahr bis (optional)", 1900, 2100, st.session_state["year_to"])
    with ff2:
        oa_only = st.checkbox("Nur Open Access (OA_URL vorhanden)", value=st.session_state["oa_only"])
    with ff3:
        sources_filter = st.multiselect(
            "Quellen einschränken (optional)",
            ["CrossRef", "OpenAlex", "OpenLibrary", "DNB_SRU", "SemanticScholar"],
            default=st.session_state["sources_filter"],
        )
    with ff4:
        publisher_contains = st.text_input(
            "Publisher/Journal enthält (optional)",
            value=st.session_state["publisher_contains"],
            placeholder="z.B. springer",
        )

    pub_types_filter = st.multiselect(
        "Publikationsarten einschränken (optional)",
        ["journal-article", "book", "book-chapter", "proceedings-article", "Buch", "Paper", "report", "thesis"],
        default=st.session_state["pub_types_filter"],
    )

    st.markdown("### D) Optional: DOI-Enrichment")
    e1, e2, e3 = st.columns(3)
    with e1:
        enable_s2 = st.checkbox("Semantic Scholar DOI-Enrichment", value=False)
    with e2:
        s2_topn = st.slider("S2 Top-N DOIs", 5, 200, 60)
    with e3:
        s2_only_missing = st.checkbox("Nur wenn Abstract/OA_URL fehlt", value=True)

    st.markdown("### E) Expertenmodus: Eigene Querylisten pro Quelle (optional)")
    expert_mode = st.checkbox(
        "Expertenmodus aktivieren (eigene Querylisten statt Auto-Query-Varianten)",
        value=False,
        help="Wenn deaktiviert, nutzt der Core automatisch Varianten aus dem Hauptsuchbegriff.",
    )

    if expert_mode:
        b1, b2 = st.columns([1, 3])
        with b1:
            if st.form_submit_button("Reset Experten-Queries auf Defaults"):
                _reset_queries_to_defaults()
                st.info("Experten-Queries wurden auf Defaults zurückgesetzt. Danach unten 'Suche starten'.")
        with b2:
            st.write("")

        q_text_crossref = st.text_area(
            "CrossRef Queries (1 pro Zeile)",
            value=st.session_state["q_text_crossref"],
            height=160,
            disabled=not enable_crossref,
        )
        q_text_openalex = st.text_area(
            "OpenAlex Queries (1 pro Zeile)",
            value=st.session_state["q_text_openalex"],
            height=160,
            disabled=not enable_openalex,
        )
        q_text_openlib = st.text_area(
            "OpenLibrary Queries (1 pro Zeile)",
            value=st.session_state["q_text_openlib"],
            height=160,
            disabled=not enable_openlib,
        )
        q_text_dnb = st.text_area(
            "DNB Queries (1 pro Zeile)",
            value=st.session_state["q_text_dnb"],
            height=160,
            disabled=not enable_dnb,
        )
    else:
        q_text_crossref = st.session_state["q_text_crossref"]
        q_text_openalex = st.session_state["q_text_openalex"]
        q_text_openlib = st.session_state["q_text_openlib"]
        q_text_dnb = st.session_state["q_text_dnb"]

    run_clicked = st.form_submit_button("Suche starten")


# Persist inputs in session_state
st.session_state["user_query"] = user_query
st.session_state["mode_label"] = mode_label
st.session_state["must_text"] = must_text
st.session_state["should_text"] = should_text
st.session_state["exclude_text"] = exclude_text
st.session_state["year_to"] = int(year_to)
st.session_state["oa_only"] = bool(oa_only)
st.session_state["sources_filter"] = sources_filter
st.session_state["pub_types_filter"] = pub_types_filter
st.session_state["publisher_contains"] = publisher_contains

if expert_mode:
    st.session_state["q_text_crossref"] = q_text_crossref
    st.session_state["q_text_openalex"] = q_text_openalex
    st.session_state["q_text_openlib"] = q_text_openlib
    st.session_state["q_text_dnb"] = q_text_dnb

# Build query lists if expert mode
q_crossref = _dedupe_lines(q_text_crossref) if (expert_mode and enable_crossref) else []
q_openalex = _dedupe_lines(q_text_openalex) if (expert_mode and enable_openalex) else []
q_openlib = _dedupe_lines(q_text_openlib) if (expert_mode and enable_openlib) else []
q_dnb = _dedupe_lines(q_text_dnb) if (expert_mode and enable_dnb) else []

st.write(
    f"Experten-Query-Anzahl: CrossRef={len(q_crossref)}, OpenAlex={len(q_openalex)}, "
    f"OpenLib={len(q_openlib)}, DNB={len(q_dnb)}"
)

# Map mode label -> mode string for core
mode = {"Offen": "open", "Normal": "normal", "Streng": "strict"}[mode_label]

must_terms = _split_csv(must_text)
should_terms = _split_csv(should_text)
exclude_terms = _split_csv(exclude_text)


def df_for_streamlit_display(df: "pd.DataFrame", n: int = 50):
    """
    Streamlit/PyArrow Workaround:
    - reset_index verhindert problematische Index-Metadaten
    - numpy/pandas scalars werden via .item() zu Python-Typen
    """
    out = df.head(n).copy()
    out = out.reset_index(drop=True)
    out.columns = [str(c) for c in out.columns]

    def _to_py(x):
        # numpy scalar -> python scalar
        if hasattr(x, "item") and not isinstance(x, (str, bytes)):
            try:
                return x.item()
            except Exception:
                return x
        return x

    # Apply element-wise
    return out.applymap(_to_py)


# Minimal guard: if no expert queries and no user_query => can't search
if run_clicked and (not expert_mode) and (not user_query.strip()):
    st.error("Bitte einen Hauptsuchbegriff eingeben oder den Expertenmodus aktivieren und Queries eintragen.")
    st.stop()

# If expert mode AND all queries empty => warn
if run_clicked and expert_mode and (len(q_crossref) + len(q_openalex) + len(q_openlib) + len(q_dnb) == 0):
    st.error("Keine Experten-Queries aktiv. Bitte mindestens eine Query eintragen oder Expertenmodus deaktivieren.")
    st.stop()


def _hash_config(cfg: dict) -> str:
    payload = json.dumps(cfg, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


if run_clicked:
    messages = st.empty()
    progress = st.progress(0)

    # Progress: Anzahl Requests ungefähr = Anzahl Queries in expert mode, sonst eine grobe Schätzung
    if expert_mode:
        steps = max(1, len(q_crossref) + len(q_openalex) + len(q_openlib) + len(q_dnb))
    else:
        # Core baut Variants; hier nur für Fortschrittsanzeige
        steps = 6  # max_variants default im Core
    done = {"value": 0}

    def progress_cb(msg: str):
        done["value"] += 1
        messages.write(msg)
        progress.progress(min(1.0, done["value"] / steps))

    config = {
        "min_year": int(min_year),
        "top_k": int(top_k),
        "rows_crossref": int(rows_crossref),
        "rows_openalex": int(rows_openalex),
        "rows_openlib": int(rows_openlib),
        "rows_dnb": int(rows_dnb),
        "enable_crossref": bool(enable_crossref),
        "enable_openalex": bool(enable_openalex),
        "enable_openlib": bool(enable_openlib),
        "enable_dnb": bool(enable_dnb),
        "use_publisher_hints": bool(use_publisher_hints),
        "enable_unpaywall": bool(enable_unpaywall),
        "enable_s2": bool(enable_s2),
        "s2_topn": int(s2_topn),
        "s2_only_if_missing": bool(s2_only_missing),

        # NEW: generic controls for core
        "mode": mode,
        "user_query": user_query.strip(),
        "must_terms": must_terms,
        "should_terms": should_terms,
        "exclude_terms": exclude_terms,
        "filters": {
            "year_to": int(year_to) if year_to else None,
            "oa_only": bool(oa_only),
            "sources": sources_filter,
            "pub_types": pub_types_filter,
            "publisher_contains": publisher_contains.strip(),
        },

        # Queries nur im Expertenmodus übergeben; sonst leer => Core nutzt Variants aus user_query
        "queries": {
            "crossref": q_crossref if expert_mode else [],
            "openalex": q_openalex if expert_mode else [],
            "openlib": q_openlib if expert_mode else [],
            "dnb": q_dnb if expert_mode else [],
        }
    }

    cfg_hash = _hash_config(config)

    same_cfg = (st.session_state["last_config_hash"] == cfg_hash)
    same_sheet = (st.session_state["last_sheet"] == sheet_name)
    same_template = (st.session_state["last_template_bytes"] == template_bytes)

    if same_cfg and same_sheet and same_template and st.session_state["last_df"] is not None:
        st.info("Gleiche Einstellungen erkannt → verwende zwischengespeichertes Ergebnis.")
        df = st.session_state["last_df"]
    else:
        with st.spinner("Suche läuft…"):
            df = run_search_to_df(
                config,
                s2_cache=st.session_state["s2_cache"],
                progress_cb=progress_cb
            )
        st.session_state["last_config_hash"] = cfg_hash
        st.session_state["last_df"] = df
        st.session_state["last_sheet"] = sheet_name
        st.session_state["last_template_bytes"] = template_bytes

    if df is None or getattr(df, "empty", True):
        st.warning("Keine Ergebnisse gefunden. Tipp: Modus auf 'Offen' stellen oder Exclude/Must lockern.")
        st.stop()

    st.success(f"Fertig! Finale Treffer: {len(df)}")
    st.dataframe(df_for_streamlit_display(df, 50), use_container_width=True)

    # In Template schreiben (Bytes)
    try:
        out_bytes = write_to_excel_bytes(df, template_bytes, sheet_name, header_row=int(header_row))
    except TypeError:
        out_bytes = write_to_excel_bytes(df, template_bytes, sheet_name)
    except Exception as e:
        st.error(
            "Fehler beim Schreiben in Excel-Template: "
            f"{e}\n\nTipp: Lade oben ein kompatibles Template herunter (mit 'Titel'-Header) "
            "oder passe Header-Zeile/Sheet an."
        )
        st.stop()

    st.download_button(
        "Excel herunterladen",
        data=out_bytes,
        file_name="literatur_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
``
